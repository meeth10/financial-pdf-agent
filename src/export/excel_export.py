"""Export a company's stored line items to an Excel workbook with one
tab per statement, plus a YoY summary tab. This is the deliverable for
the 'send me the information in a separate tab' requirement.
"""

import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

STATEMENT_SHEETS = [
    ("balance_sheet", "Balance Sheet"),
    ("income_statement", "Income Statement"),
    ("cash_flow", "Cash Flow"),
]


def _write_statement_sheet(wb: Workbook, conn: sqlite3.Connection, entity: str,
                            statement_key: str, sheet_name: str) -> None:
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """SELECT DISTINCT metric, metric_raw FROM line_items
           WHERE entity = ? AND statement = ? ORDER BY metric""",
        (entity, statement_key),
    ).fetchall()
    periods = [r["period"] for r in conn.execute(
        "SELECT DISTINCT period FROM line_items WHERE entity = ? AND statement = ? ORDER BY period",
        (entity, statement_key),
    ).fetchall()]

    ws = wb.create_sheet(sheet_name)
    ws.append(["Metric"] + periods)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row in rows:
        line = [row["metric_raw"] or row["metric"]]
        for period in periods:
            val = conn.execute(
                """SELECT value FROM line_items
                   WHERE entity = ? AND statement = ? AND metric = ? AND period = ?
                   LIMIT 1""",
                (entity, statement_key, row["metric"], period),
            ).fetchone()
            line.append(val["value"] if val else None)
        ws.append(line)

    for i, _ in enumerate([("Metric",)] + periods, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 22


def _write_yoy_sheet(wb: Workbook, conn: sqlite3.Connection, entity: str) -> None:
    conn.row_factory = sqlite3.Row
    ws = wb.create_sheet("YoY Summary")
    ws.append(["Metric", "Statement", "Period A", "Period B", "Change", "Change %"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    metrics = conn.execute(
        "SELECT DISTINCT metric, statement FROM line_items WHERE entity = ?", (entity,)
    ).fetchall()
    for m in metrics:
        periods = [r["period"] for r in conn.execute(
            """SELECT DISTINCT period FROM line_items
               WHERE entity = ? AND metric = ? AND statement = ? ORDER BY period""",
            (entity, m["metric"], m["statement"]),
        ).fetchall()]
        for a, b in zip(periods, periods[1:]):
            va = conn.execute(
                "SELECT value FROM line_items WHERE entity=? AND metric=? AND statement=? AND period=?",
                (entity, m["metric"], m["statement"], a)).fetchone()
            vb = conn.execute(
                "SELECT value FROM line_items WHERE entity=? AND metric=? AND statement=? AND period=?",
                (entity, m["metric"], m["statement"], b)).fetchone()
            if va and vb and va["value"] not in (None, 0):
                change = vb["value"] - va["value"]
                pct = change / va["value"] * 100
                ws.append([m["metric"], m["statement"], a, b, round(change, 2), round(pct, 1)])

    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(i)].width = 18


def export_entity(conn: sqlite3.Connection, entity: str, output_path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet
    for key, name in STATEMENT_SHEETS:
        _write_statement_sheet(wb, conn, entity, key, name)
    _write_yoy_sheet(wb, conn, entity)
    wb.save(output_path)
    return output_path
